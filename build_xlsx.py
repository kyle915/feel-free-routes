"""Build the internal Excel route workbook from schedule.json."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open("schedule.json"))
P = data["program"]; MK = data["markets"]; SCH = data["schedule"]

INK = "1A1A1A"; WHITE = "FFFFFF"; LIGHT = "F2F2F2"; MUTE = "6B6B6B"
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def hdr(ws, row, cols, fill):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = center; cell.border = border

wb = Workbook()
ws = wb.active; ws.title = "Overview"; ws.sheet_view.showGridLines = False
ws["A1"] = "FEEL FREE — SUMMER 2026 GUERRILLA FIELD SAMPLING"
ws["A1"].font = Font(bold=True, size=16, color=INK)
ws["A2"] = "Weekly Route Plan  |  Internal Ops Master  |  Ignite Productions"
ws["A2"].font = Font(size=11, italic=True, color=MUTE)
def _mk_label(name, m):
    if m.get("tbd"):
        return f"{name} (TBD)"
    if "concluded" in m.get("status", ""):
        return f"{name} (concluded)"
    return name

facts = [
    ("Client", P["client"]),
    ("Program window", f'{P["start_date"]}  to  {P["end_date"]}  (varies by market — see below)'),
    ("Cadence", P["cadence"] + "  (4 shifts/market/week while active)"),
    ("Markets", " · ".join(_mk_label(name, m) for name, m in MK.items())),
    ("Shift structure", f'{P["bas_per_shift"]} BAs/shift · {P["billable_hours_per_ba"]} billable hrs/BA'),
    ("Thu / Fri window", f'Call {P["windows"]["Thu"]["call"]} · Active {P["windows"]["Thu"]["active"]}'),
    ("Sat / Sun window", f'Call {P["windows"]["Sat"]["call"]} · Active {P["windows"]["Sat"]["active"]}'),
    ("Total shifts (this plan)", f'{len(SCH)} shifts · {len(SCH)*2} BA-slots · {len(SCH)*5:.0f} BA-hours'),
    ("Activation format", "Mobile street/sidewalk sampling. Start at warehouse (product pickup), then rotate route stops."),
]
r = 4
for k, v in facts:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10, color=INK)
    cc = ws.cell(row=r, column=2, value=v); cc.font = Font(size=10, color=INK); cc.alignment = wrap
    r += 1
r += 1
ws.cell(row=r, column=1, value="COMPLIANCE (read before every shift)").font = Font(bold=True, size=11, color="C0392B")
r += 1
ws.cell(row=r, column=1, value=P["compliance"]).font = Font(size=10, color=INK)
ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=6)
ws.cell(row=r, column=1).alignment = wrap
r += 5
ws.cell(row=r, column=1, value="SHIFTS BY MARKET").font = Font(bold=True, size=11, color=INK)
r += 1
hdr(ws, r, ["Market", "Code", "Warehouse (shift start)", "Shifts", "BA-Hours", "Product"], INK)
r += 1
from collections import Counter
cnt = Counter(s["market"] for s in SCH)
for market, m in MK.items():
    n = cnt[market]
    vals = [market, m["code"], m["warehouse"], n, f"{n*5:.0f}", m["product"]]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = Font(size=10, color=INK); cell.alignment = wrap; cell.border = border
        if j == 2:
            cell.fill = PatternFill("solid", fgColor=m["color"].lstrip("#"))
            cell.font = Font(size=10, bold=True, color=WHITE); cell.alignment = center
    r += 1
for i, w in enumerate([22, 8, 46, 9, 11, 34], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wm = wb.create_sheet("Master Schedule"); wm.sheet_view.showGridLines = False
cols = ["Week", "Date", "Day", "Market", "Corridor", "Call", "Active Window",
        "Stop 1 (zone)", "Stop 1 time", "Stop 1 address",
        "Stop 2 (zone)", "Stop 2 time", "Stop 2 address", "BAs", "Hrs", "Event / Note"]
hdr(wm, 1, cols, INK)
r = 2
for s in SCH:
    st = s["stops"]
    s1 = st[0] if len(st) > 0 else {"zone": "", "time": "", "address": ""}
    s2 = st[1] if len(st) > 1 else {"zone": "", "time": "", "address": ""}
    row = [s["week"], s["date_pretty"], s["day"], s["market"], s["corridor"], s["call"], s["active"],
           s1["zone"], s1["time"], s1["address"], s2["zone"], s2["time"], s2["address"],
           s["bas"], s["hours"], s["event"] or ""]
    for j, v in enumerate(row, 1):
        cell = wm.cell(row=r, column=j, value=v)
        cell.font = Font(size=9, color=INK); cell.alignment = wrap; cell.border = border
        if j == 4:
            cell.fill = PatternFill("solid", fgColor=s["color"].lstrip("#")); cell.font = Font(size=9, bold=True, color=WHITE)
        if s["event"]:
            wm.cell(row=r, column=16).fill = PatternFill("solid", fgColor="FFF4D6")
    r += 1
wm.freeze_panes = "A2"; wm.auto_filter.ref = f"A1:P{r-1}"
for i, w in enumerate([6, 11, 6, 18, 24, 9, 14, 22, 12, 24, 22, 12, 24, 5, 5, 26], 1):
    wm.column_dimensions[get_column_letter(i)].width = w

for market, m in MK.items():
    wsx = wb.create_sheet(m["code"]); wsx.sheet_view.showGridLines = False
    wsx["A1"] = f"{market}  —  Weekly Route Plan"
    wsx["A1"].font = Font(bold=True, size=14, color=m["color"].lstrip("#"))
    wsx["A2"] = f'Warehouse (shift start): {m["warehouse"]}'; wsx["A2"].font = Font(size=10, italic=True, color=MUTE)
    wsx["A3"] = f'Product: {m["product"]}   |   2 BAs/shift · 5 hrs/BA'; wsx["A3"].font = Font(size=10, color=MUTE)
    wsx["A4"] = m.get("status", ""); wsx["A4"].font = Font(size=10, bold=True, color=INK)
    hdr(wsx, 5, ["Wk", "Date", "Day", "Active Window", "Corridor", "Vibe",
                 "Stop 1 (zone · time · address)", "Stop 2 (zone · time · address)",
                 "Why this stop", "Event / Note"], m["color"].lstrip("#"))
    r = 6
    mkt_rows = [x for x in SCH if x["market"] == market]
    if not mkt_rows:
        wsx.cell(row=r, column=1, value="Schedule TBD — corridors and dates to be announced.").font = (
            Font(size=11, italic=True, color=MUTE))
        wsx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    for s in mkt_rows:
        st = s["stops"]
        fmt = lambda x: f'{x["zone"]}\n{x["time"]}\n{x["address"]}'
        row = [s["week"], s["date_pretty"], s["day"], s["active"], s["corridor"], s["vibe"],
               fmt(st[0]) if len(st) > 0 else "", fmt(st[1]) if len(st) > 1 else "",
               s["why"], s["event"] or ""]
        for j, v in enumerate(row, 1):
            cell = wsx.cell(row=r, column=j, value=v)
            cell.font = Font(size=9, color=INK); cell.alignment = wrap; cell.border = border
            if r % 2 == 0: cell.fill = PatternFill("solid", fgColor=LIGHT)
            if s["event"] and j == 10: cell.fill = PatternFill("solid", fgColor="FFF4D6")
        r += 1
    wsx.freeze_panes = "A6"
    for i, w in enumerate([5, 11, 6, 14, 22, 22, 30, 30, 34, 22], 1):
        wsx.column_dimensions[get_column_letter(i)].width = w

wb.save("Feel_Free_Summer2026_Route_Plan_INTERNAL.xlsx")
print("Saved workbook:", wb.sheetnames)
